"""
Diagnose why the dashboard shows no utilization.

Run it the same way you run the generator:

    python check_workbook.py

It reads the workbook, reproduces exactly what generate_report.py does to it,
and reports where the data stops flowing. It changes nothing.
"""

import re
import sys
import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is not installed.  pip install openpyxl")

HERE = Path(__file__).resolve().parent
CANDIDATES = [
    HERE.parent / "03_Output files" / "Employee Details.xlsx",
    HERE / "Employee Details.xlsx",
    HERE.parent / "01_Source" / "Employee Details.xlsx",
]

UTIL_SHEETS = {
    "Utilization Full_Jul_Jun", "Utilization Full_May_April", "Utilization Full_Apr_Mar",
    "Employee Utilization_Jul_Jun", "Employee Utilization_May_April",
    "Employee Utilization_Apr_Mar", "Employee Monthly Utilization",
}
KNOWN = UTIL_SHEETS | {
    "Employee Details", "Employee Skills", "Skill Mapping",
    "Employee Skills Hierarchy", "Hourly Rates",
}
MON = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def flat(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def find(headers, *names):
    """Locate a column by any of several spellings, ignoring case/spacing."""
    idx = {flat(h): h for h in headers}
    for n in names:
        if flat(n) in idx:
            return idx[flat(n)]
    return None


def month_key(raw):
    """The same parse the dashboard does: anything -> 'YYYY-MM' or ''."""
    if raw is None or raw == "":
        return ""
    if hasattr(raw, "year"):
        return f"{raw.year}-{raw.month:02d}"
    s = str(raw).strip()
    m = re.match(r"^(\d{4})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"([A-Za-z]{3,})[\s\-/]+(\d{4})", s)
    if m and m.group(1)[:3].lower() in MON:
        return f"{m.group(2)}-{MON.index(m.group(1)[:3].lower()) + 1:02d}"
    m = re.match(r"^(\d{1,2})[-/](\d{4})$", s)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    return ""


def rows_of(ws):
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return [], []
    headers = [("" if h is None else str(h).strip()) for h in data[0]]
    out = []
    for raw in data[1:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
            continue
        out.append({h: (raw[i] if i < len(raw) else None) for i, h in enumerate(headers)})
    return headers, out


def num(v):
    """'79.46%' -> 79.46, '1,234.5' -> 1234.5. Matches what the generator does."""
    if v is None or v == "" or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if s.endswith("%"):
        s = s[:-1].strip()
    try:
        return float(s)
    except ValueError:
        return None


def norm_id(v):
    """How the dashboard compares IDs: as a string. 101351300.0 and 101351300
    are different strings, which is a common and invisible join failure."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main():
    path = next((p for p in CANDIDATES if p.exists()), None)
    if path is None:
        sys.exit("Could not find 'Employee Details.xlsx' in the usual places:\n  "
                 + "\n  ".join(str(p) for p in CANDIDATES))
    print(f"Workbook: {path}\n")
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---------- sheets ----------
    print("=" * 66)
    print("SHEETS")
    print("=" * 66)
    for t in wb.sheetnames:
        mark = "read " if t.strip() in KNOWN else "IGNORED (name not recognised)"
        print(f"  {t:<34} {mark}")
    if "Hourly Rates" not in wb.sheetnames:
        print("\n  !! No 'Hourly Rates' sheet -> the Rate Analysis tab will not appear.")

    # ---------- roster ----------
    det_name = next((t for t in wb.sheetnames if t.strip() == "Employee Details"), wb.sheetnames[0])
    dh, det = rows_of(wb[det_name])
    id_col = find(dh, "WorkdayID", "Workday ID", "Workday Id")
    if id_col is None:
        sys.exit(f"\n'{det_name}' has no WorkdayID column. Found: {dh}")
    roster = {norm_id(e[id_col]) for e in det if norm_id(e[id_col])}
    print(f"\n  Roster: {len(roster)} people (ID column '{id_col}')")

    # ---------- each utilization sheet ----------
    any_ok = False
    for t in wb.sheetnames:
        if t.strip() not in UTIL_SHEETS:
            continue
        print("\n" + "=" * 66)
        print(f"SHEET: {t}")
        print("=" * 66)
        h, rows = rows_of(wb[t])
        print(f"  rows: {len(rows)}")
        print(f"  columns: {h}")

        c_id = find(h, "Workday ID", "WorkdayID", "Workday Id")
        c_mo = find(h, "EoM", "EOM", "Month Year", "MonthYear", "Month")
        c_ut = find(h, "Utilisation%", "Utilization%", "Utilisation", "Utilization")
        c_ch = find(h, "Chargeable Hours", "ChargeableHours", "Charged Hours")
        c_sd = find(h, "Standard Hours", "StandardHours", "Std Hours")
        for label, col in [("Workday ID", c_id), ("EoM / Month", c_mo),
                           ("Utilisation%", c_ut), ("Chargeable Hours", c_ch),
                           ("Standard Hours", c_sd)]:
            print(f"    {label:<18} -> {col if col else '*** NOT FOUND ***'}")
        if not c_id or not c_mo:
            print("  !! Without both an ID and a month column this sheet is dropped entirely.")
            continue

        sample = rows[0]
        print(f"\n  first row: ID={sample[c_id]!r}  month={sample[c_mo]!r} "
              f"({type(sample[c_mo]).__name__})")

        good_m = [r for r in rows if month_key(r[c_mo])]
        months = sorted({month_key(r[c_mo]) for r in good_m})
        print(f"  months parsed: {len(good_m)}/{len(rows)} rows -> {len(months)} distinct")
        if months:
            print(f"    {months[0]} .. {months[-1]}")
        if len(good_m) < len(rows):
            bad = [r[c_mo] for r in rows if not month_key(r[c_mo])][:5]
            print(f"    !! unparseable month values, e.g. {bad}")

        matched = [r for r in good_m if norm_id(r[c_id]) in roster]
        print(f"  IDs matching the roster: {len(matched)}/{len(good_m)}")
        if good_m and not matched:
            u = sorted({norm_id(r[c_id]) for r in good_m})[:4]
            rr = sorted(roster)[:4]
            print("    !! NOTHING JOINS. The dashboard shows no utilization.")
            print(f"       sheet IDs : {u}")
            print(f"       roster IDs: {rr}")
            print("       Compare them carefully - trailing '.0', spaces or text-vs-number")
            print("       make two identical-looking IDs different to the dashboard.")
        elif len(matched) < len(good_m):
            extra = sorted({norm_id(r[c_id]) for r in good_m} - roster)
            print(f"    {len(extra)} IDs are not on the roster - these read as leavers: {extra[:5]}")

        # Duplicate person-months: the dashboard sums them, but if they are not
        # meant to be there they will double-count.
        seen = {}
        for r in matched:
            k = (norm_id(r[c_id]), month_key(r[c_mo]))
            seen[k] = seen.get(k, 0) + 1
        dup = {k: v for k, v in seen.items() if v > 1}
        print(f"  Person-months: {len(seen)} distinct from {len(matched)} rows")
        if dup:
            print(f"    !! {len(dup)} person-months appear on more than one row "
                  f"(up to {max(dup.values())}).")
            print("       The dashboard sums them. If that is a split by BU or project this is")
            print("       correct; if the rows are accidental repeats, hours will be doubled.")
            for k, v in list(dup.items())[:3]:
                print(f"       e.g. {k[0]} / {k[1]} appears {v} times")

        if c_sd:
            # Average utilization is total chargeable / total standard, so rows
            # missing standard hours are left out of the ratio entirely. A
            # partially-filled column quietly narrows what the average covers.
            have = [r for r in matched if num(r[c_sd]) not in (None, 0)]
            print(f"  Standard Hours present on {len(have)}/{len(matched)} joined rows")
            if matched and len(have) < len(matched):
                print("    !! Average utilization is computed only over the rows that have")
                print("       standard hours. The rest contribute nothing to the ratio.")
            if have and c_ch:
                ch = sum(num(r[c_ch]) or 0 for r in have)
                sd = sum(num(r[c_sd]) or 0 for r in have)
                if sd:
                    print(f"  => average utilization = {ch:,.1f} / {sd:,.1f} = {ch / sd * 100:.1f}%")

        if c_ut:
            vals = [num(r[c_ut]) for r in matched]
            vals = [v for v in vals if isinstance(v, (int, float))]
            if vals:
                lo, hi = min(vals), max(vals)
                print(f"  Utilisation% range: {lo} .. {hi}")
                if hi <= 2:
                    print("    !! These look like fractions (0.85), not percentages (85).")
                    print("       The dashboard reads them literally, so everyone shows ~1%.")
            else:
                print("  !! Utilisation% has no numeric values.")
        if matched:
            any_ok = True

    print("\n" + "=" * 66)
    print("VERDICT")
    print("=" * 66)
    print("  Utilization will render." if any_ok else
          "  No utilization sheet produced usable, joinable rows - Pulse, Team\n"
          "  Analytics and Rate Analysis will all be empty. See the !! lines above.")


if __name__ == "__main__":
    main()
