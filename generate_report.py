#!/usr/bin/env python3
"""
Employee Bio Dashboard - PwC-branded, single-file HTML.

Reads:
  Employee Details.xlsx   (sheets: Employee Details / Employee Skills /
                           Employee Monthly Utilization)
  Images/<PhotoID>.png    (photos referenced by the PhotoID column)
  template.html           (page template, lives next to this script)

Writes:
  Employee_Dashboard.html (self-contained, images base64-embedded,
                           shareable with clients - works offline)

Reusable: keep the schema the same, replace the .xlsx and/or add photos,
then rerun:
    python generate_report.py
"""

import argparse
import base64
import json
import datetime
import re
import struct
import zlib
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl

# Optional: Pillow for photo downscaling. Massive HTML-size win at scale.
# pip install Pillow
try:
    from PIL import Image, ImageOps
    HAVE_PIL = True
except ImportError:
    HAVE_PIL = False

# Maximum dimension (px) for embedded photos. 240 covers both the directory
# card and the bio modal at high-DPI quality while keeping file size tiny.
PHOTO_MAX_DIM = 320
PHOTO_JPEG_QUALITY = 72

# ---- Configuration -------------------------------------------------------
SCRIPT_DIR   = Path(__file__).resolve().parent          # 02_Scripts & ETL
ROOT         = SCRIPT_DIR.parent                         # Deals Skills and Bio
SOURCE_DIR   = ROOT / "01_Source"                        # raw input files + Images + logo
OUTPUT_DIR   = ROOT / "03_Output files"                  # generated HTML
OUTPUT_DIR.mkdir(exist_ok=True)
INPUT_FILE   = OUTPUT_DIR / "Employee Details.xlsx"      # master file (built by your ETL) lives with the outputs
IMAGES_DIR   = SOURCE_DIR / "Images"
TEMPLATE     = SCRIPT_DIR / "template.html"
MOBILE_TEMPLATE = SCRIPT_DIR / "template_mobile.html"   # phone-first companion build
OUTPUT_FILE  = OUTPUT_DIR / "Employee_Dashboard.html"

# PwC logo lives in the working folder. Looked up by stem in this order.
LOGO_CANDIDATES = ["PwC Logo", "pwc logo", "PwC_Logo", "logo"]

PHOTO_EXTS   = [".png", ".jpg", ".jpeg", ".webp", ".gif"]
MIME = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".webp": "image/webp", ".gif": "image/gif"}

# Two columns in the source share the name "Experience". Rename the second
# one (years at PwC) for clarity.
EXP_TOTAL = "Experience"
EXP_PWC   = "PwC Experience"

# Role grade codes -> full display names. Applied at load; a no-op if the
# workbook already uses full names. Keeps display/filters/grouping consistent.
ROLE_RENAME = {
    "md": "Managing Director", "managing director": "Managing Director",
    "d": "Director", "director": "Director",
    "sm": "Senior Manager", "sr manager": "Senior Manager", "senior manager": "Senior Manager",
    "m": "Manager", "manager": "Manager",
    "sa": "Senior Associate", "sa1": "Senior Associate", "sa2": "Senior Associate",
    "sa3": "Senior Associate", "senior associate": "Senior Associate",
    "a2": "Associate 2", "associate 2": "Associate 2",
    "a1": "Associate", "a": "Associate", "a3": "Associate", "associate": "Associate",
}


def rename_role(v):
    return ROLE_RENAME.get(str(v or "").strip().lower(), v)


def normalise_headers(raw):
    """De-duplicate header names. Second 'Experience' becomes 'PwC Experience'."""
    seen, out = {}, []
    for h in raw:
        h = "" if h is None else str(h).strip()
        if h in seen:
            seen[h] += 1
            out.append(EXP_PWC if h == "Experience" else f"{h} ({seen[h]})")
        else:
            seen[h] = 1
            out.append(h)
    return out


def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = normalise_headers(rows[0])
    records = []
    for raw in rows[1:]:
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in raw):
            continue
        rec = {}
        for i, header in enumerate(headers):
            v = raw[i] if i < len(raw) else None
            if isinstance(v, datetime.datetime):
                v = v.strftime("%Y-%m-%d")
            elif isinstance(v, datetime.date):
                v = v.isoformat()
            rec[header] = "" if v is None else v
        records.append(rec)
    return records


def _recover_xlsx_bytes(raw):
    """Rebuild a clean .xlsx when its zip central directory / end-of-archive
    record is missing. This can happen when the file is read mid-save or the
    sync layer truncates the tail. We scan the intact local file headers at the
    front of the archive and re-zip every member we can fully decompress."""
    out = BytesIO()
    i, sig = 0, b"PK\x03\x04"
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        while True:
            j = raw.find(sig, i)
            if j < 0:
                break
            try:
                (_ver, _flag, method, _mt, _md, _crc, csize, usize,
                 fnlen, eflen) = struct.unpack("<HHHHHIIIHH", raw[j + 4:j + 30])
            except struct.error:
                break
            name = raw[j + 30:j + 30 + fnlen].decode("utf-8", "replace")
            start = j + 30 + fnlen + eflen
            comp = raw[start:start + csize] if csize else b""
            i = j + 4
            try:
                data = zlib.decompress(comp, -15) if method == 8 else comp
            except Exception:
                continue
            if usize and len(data) != usize:   # this member was itself truncated
                continue
            z.writestr(name, data)
    out.seek(0)
    return out


def load_workbook_data(path):
    if not path.exists():
        raise FileNotFoundError(f"Could not find input file: {path}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[warn] Workbook archive looks truncated ({e}); recovering from local headers...")
        wb = openpyxl.load_workbook(_recover_xlsx_bytes(path.read_bytes()), data_only=True)
    keymap = {
        "Employee Details": "details",
        "Employee Skills": "skills",
        "Skill Mapping": "skills",          # real-file sheet name
        "Employee Skills Hierarchy": "skills",  # renamed sheet (adds Skill Group column)
        "Employee Monthly Utilization": "utilization",
        "Employee Utilization_Jul_Jun": "util_jj",   # FY Jul-Jun view (Pulse toggle)
        "Employee Utilization_May_April": "util_am", # Performance Year May-Apr view (Pulse toggle)
        "Employee Utilization_Apr_Mar": "util_am",   # legacy tab name, kept as fallback
        # Full sheets: every person who worked in the period, not just those still
        # on the roster. Carries Standard Hours and the grade/competency/territory
        # held in that month, so history is attributed to where someone actually was.
        "Utilization Full_Jul_Jun":   "full_jj",
        "Utilization Full_May_April": "full_am",
        "Utilization Full_Apr_Mar":   "full_am",     # legacy tab name, kept as fallback
        "Hourly Rates": "rates",                     # Role x Territory rate card
    }
    data = {"details": [], "skills": [], "utilization": [], "util_jj": [], "util_am": [],
            "full_jj": [], "full_am": [], "rates": []}
    for ws in wb.worksheets:
        key = keymap.get(ws.title.strip())
        if key is None and not data["details"]:
            key = "details"  # fall back: first sheet is details
        if key:
            data[key] = read_sheet(ws)
    # Fallbacks so the dashboard works whether or not the FY-specific sheets are
    # present. If only the legacy monthly sheet exists, use it for both FY views;
    # if only FY sheets exist, use Jul-Jun as the global utilization source.
    if not data["utilization"]:
        data["utilization"] = data["util_jj"] or data["util_am"]
    if not data["util_jj"]:
        data["util_jj"] = data["utilization"]
    if not data["util_am"]:
        data["util_am"] = data["utilization"]
    # The Full sheets supersede the older ones where present. They are normalised
    # to the same key names the dashboard already reads, plus the extra columns,
    # so nothing downstream has to know which sheet it came from.
    full_jj = normalise_full(data["full_jj"], "Utilization Full_Jul_Jun") if data["full_jj"] else []
    full_am = normalise_full(data["full_am"], "Utilization Full_May_April") if data["full_am"] else []
    # Only supersede when the Full sheet actually parsed. A sheet that is present
    # but unreadable must not replace working data with nothing — that empties
    # every utilization view and silently removes the Rate Analysis tab.
    # Canonicalise the join key on both sides before anything downstream sees it.
    for e in data["details"]:
        for k in ("WorkdayID", "Workday ID"):
            if k in e:
                e[k] = canon_id(e[k])
    for key in ("utilization", "util_jj", "util_am"):
        for r in data[key]:
            if "WorkdayID" in r:
                r["WorkdayID"] = canon_id(r["WorkdayID"])

    if full_jj:
        data["util_jj"] = full_jj
        data["utilization"] = full_jj
    if full_am:
        data["util_am"] = full_am

    # ---- tell the operator what was and was not picked up -------------------
    seen = sorted({ws.title.strip() for ws in wb.worksheets})
    matched = sorted({t for t in seen if t in keymap})
    print(f"[info] Sheets read: {', '.join(matched) if matched else '(none matched by name)'}")
    ignored = [t for t in seen if t not in keymap]
    if ignored:
        print(f"[info] Sheets ignored (name not recognised): {', '.join(ignored)}")
    if not data["rates"]:
        print("[warn] No 'Hourly Rates' sheet found - the RATE ANALYSIS TAB WILL NOT APPEAR.")
        print("[warn]   Add a sheet named exactly 'Hourly Rates' with columns:")
        print("[warn]   Role | Territory | Hourly Rate (USD).  Use Territory = 'Standard' for the default rate.")
    if not data["utilization"]:
        print("[warn] No usable utilization rows - Pulse, Team Analytics and Rate Analysis will be empty.")
    return data


_MON = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def build_leavers(details, *util_sets):
    """Employee-shaped records for people who worked in a period but are no longer
    on the roster.

    There is no status column to read, so 'left' is defined structurally: present
    in a utilization sheet, absent from Employee Details. That cannot drift out of
    step the way a flag can. Attributes are taken from their most recent month, so
    a leaver is described as they were when they left."""
    roster = {str(e.get("WorkdayID", "")).strip() for e in details}
    latest = {}
    for rows in util_sets:
        for r in rows:
            wid = str(r.get("WorkdayID", "")).strip()
            if not wid or wid in roster:
                continue
            key = _month_sort_key(r.get("Month Year"))
            if wid not in latest or key > latest[wid][0]:
                latest[wid] = (key, r)
    out = []
    for wid, (_, r) in sorted(latest.items()):
        out.append({
            "WorkdayID":         r.get("WorkdayID"),
            "Name":              r.get("Name"),
            "Employee ID":       r.get("Employee ID"),
            "Role":              r.get("Role"),
            "Competency Group":  r.get("Competency Group"),
            "Competency":        r.get("Competency"),
            "Competency Filter": r.get("Competency Filter"),
            "Territory Group":   r.get("Territory Group"),
            "Territory":         r.get("Territory"),
            "Territory Filter":  r.get("Territory Filter"),
            "Status":            "Left",
            "_leaver":           True,
            # the avatar fallback the dashboard draws when there is no photo
            "_initials":         "".join(w[0] for w in str(r.get("Name") or "").split()[:2]).upper(),
        })
    return out


def _month_sort_key(my):
    """'Mar-2026' -> (2026, 3). Unparseable months sort first so they never win
    the 'most recent row' comparison."""
    try:
        mo, yr = str(my).split("-")
        return (int(yr), _MON.index(mo[:3].title()) + 1)
    except Exception:
        return (0, 0)


def num_val(v):
    """Coerce a spreadsheet value to a number.

    Real extracts often store these columns as text: '79.46%', '1,234.5',
    ' 168 '. JavaScript's Number() returns NaN for all of them, so the row is
    read as 'nothing reported' and the figure silently disappears. Parsing here
    means every consumer downstream gets a plain number.

    A percent sign is stripped, not rescaled: '79.46%' is 79.46, because that is
    what the column means. A genuine Excel percentage *cell* arrives as 0.7946
    and is left alone — rescaling that would be a guess, so it is reported
    instead."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if s.endswith("%"):
        s = s[:-1].strip()
    try:
        return float(s)
    except ValueError:
        return None


def canon_id(v):
    """Canonical Workday ID.

    The dashboard joins people to hours with a plain string compare, so
    ' 101351303', '101351303' and '101351303.0' are three different people to
    it — and the symptom is simply that all utilization disappears. Excel
    produces all three readily. Normalising here fixes every downstream join at
    once."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s


def _pick(row, *names):
    """Fetch a column by any of several spellings, ignoring case, spaces and
    punctuation. Real extracts vary — 'Workday ID' / 'WorkdayID' / 'WORKDAY_ID'
    are the same column, and a silent miss here empties the whole sheet."""
    if not hasattr(_pick, "_cache"):
        _pick._cache = {}
    key = id(row)
    idx = _pick._cache.get(key)
    if idx is None:
        idx = {re.sub(r"[^a-z0-9]", "", str(k).lower()): k for k in row}
        _pick._cache = {key: idx}          # single-entry cache; rows share headers
    for n in names:
        k = idx.get(re.sub(r"[^a-z0-9]", "", n.lower()))
        if k is not None and row.get(k) not in (None, ""):
            return row[k]
    return None


def normalise_full(rows, label=""):
    """Map a 'Utilization Full_*' sheet onto the key names the dashboard reads.

    EoM is an end-of-month date; everything downstream expects 'Mon-YYYY', so it
    is converted here rather than teaching the template a second date format.
    Rows without a resolvable month or Workday ID are dropped, since they cannot
    be joined to anything."""
    out = []
    dropped_month = dropped_id = 0
    for r in rows:
        eom = _pick(r, "EoM", "EOM", "Month Year", "MonthYear", "Month")
        if hasattr(eom, "year"):
            month = f"{_MON[eom.month - 1]}-{eom.year}"
        else:                                    # already a string, or unusable
            month = str(eom or "").strip()
            if not month:
                dropped_month += 1
                continue
        wid = _pick(r, "Workday ID", "WorkdayID", "Workday Id")
        if wid in (None, ""):
            dropped_id += 1
            continue
        out.append({
            "WorkdayID":        canon_id(wid),
            "Name":             _pick(r, "EMP Name", "Name", "Employee Name"),
            "Employee ID":      _pick(r, "EMP ID", "Employee ID", "EmpID"),
            "Month Year":       month,
            "Chargeable Hours": num_val(_pick(r, "Chargeable Hours", "ChargeableHours", "Charged Hours")),
            "Training Hours":   num_val(_pick(r, "Training Hours", "TrainingHours")),
            "Utilization":      num_val(_pick(r, "Utilisation%", "Utilization%", "Utilisation", "Utilization")),
            # new, and the reason the Full sheets are worth switching to
            "Standard Hours":   num_val(_pick(r, "Standard Hours", "StandardHours", "Std Hours")),
            "Role":             _pick(r, "EMP Designation", "Designation", "Role", "Grade"),
            "Competency Group": _pick(r, "Competency Group"),
            "Competency":       _pick(r, "Competency"),
            "Competency Filter":_pick(r, "Competency Filter"),
            "Territory Group":  _pick(r, "Territory Group"),
            "Territory":        _pick(r, "Territory"),
            "Territory Filter": _pick(r, "Territory Filter"),
        })
    # A person-month arriving as several rows is legitimate — a split by BU, OU
    # or project — and the dashboard sums them. Worth reporting either way, so
    # the totals are never a surprise.
    seen = {}
    for r in out:
        k = (str(r["WorkdayID"]), r["Month Year"])
        seen[k] = seen.get(k, 0) + 1
    dup = {k: v for k, v in seen.items() if v > 1}
    if dup:
        worst = max(dup.values())
        print(f"[info] '{label}': {len(out)} rows cover {len(seen)} person-months; "
              f"{len(dup)} of them arrive as multiple rows (up to {worst}). Hours are summed.")
    if rows and not out:
        print(f"[warn] '{label}' has {len(rows)} rows but none were usable "
              f"(no month: {dropped_month}, no Workday ID: {dropped_id}).")
        print(f"[warn]   columns found: {list(rows[0].keys())}")
        print("[warn]   expected at least 'Workday ID' and 'EoM'. This sheet is being ignored.")
    return out


def _encode_one(path):
    """Embed the image exactly as it is in the folder, with no re-encoding,
    so quality is preserved. Sizing/compression is handled upstream by the
    user's own image script, so we must not compress it again here."""
    ext = path.suffix.lower()
    mime = MIME.get(ext, "image/png")
    raw = path.read_bytes()
    return (f"data:{mime};base64," + base64.b64encode(raw).decode("ascii"), len(raw))


def _norm_id(v):
    """Normalise an id (number or text) to a plain string for matching.
    Handles int/float (534034.0 -> '534034') and stray whitespace."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def load_photos(images_dir):
    photos = {}
    if not images_dir.exists():
        return photos
    total_bytes = 0
    for f in images_dir.iterdir():
        if f.is_file() and f.suffix.lower() in PHOTO_EXTS:
            uri, size = _encode_one(f)
            total_bytes += size
            stem = f.stem.strip()
            photos[_norm_id(stem)] = uri   # normalised (matches numeric Employee ID)
            photos[stem] = uri             # also keep the raw filename stem
    if photos:
        avg = total_bytes / len(set(photos.values())) if photos else 0
        mode = "embedded as-is (original quality)"
        print(f"[ok] Photos {mode}: {len(set(photos.values()))} files, "
              f"total {total_bytes/1024:.0f} KB, avg {avg/1024:.0f} KB/photo")
    return photos


def photo_for(rec, photos):
    # Photos are matched by the image filename stem. Real data names images by
    # Employee ID, so try that first; fall back to PhotoID / WorkdayID.
    for key in ("Employee ID", "PhotoID", "WorkdayID"):
        nid = _norm_id(rec.get(key, ""))
        if not nid:
            continue
        uri = photos.get(nid, "")
        if uri:
            return uri
    return ""


def load_logo():
    """Find the PwC logo in the working folder and return a base64 data URI."""
    for stem in LOGO_CANDIDATES:
        for ext in PHOTO_EXTS:
            p = SOURCE_DIR / f"{stem}{ext}"
            if p.exists():
                b64 = base64.b64encode(p.read_bytes()).decode("ascii")
                return f"data:{MIME[ext.lower()]};base64,{b64}"
    return ""


def initials(name):
    parts = [p for p in str(name or "").split() if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[-1][0]).upper()


# Address-PIN geocoding used to run here, feeding Latitude/Longitude into the
# employee JSON for the Locations map. That tab is parked (see
# 05_Parked features/) and no template reads those columns, so the lookup was
# removed rather than left to load 19,550 PINs on every run. The code is kept
# in 05_Parked features/geocoding_PARKED.py next to the tab it belongs to.


def main():
    parser = argparse.ArgumentParser(description="Build the employee dashboard HTML.")
    parser.add_argument("--pbi-url", default="",
                        help="Optional Power BI report URL linked from bios in the restricted build")
    args = parser.parse_args()

    data = load_workbook_data(INPUT_FILE)
    for _r in data["details"]:                     # normalise Role grade codes
        _r["Role"] = rename_role(_r.get("Role"))
    photos = load_photos(IMAGES_DIR)
    logo_uri = load_logo()

    enriched = []
    for r in data["details"]:
        c = dict(r)
        c["_photo"] = photo_for(r, photos)
        c["_initials"] = initials(r.get("Name"))
        enriched.append(c)

    # Photo-match diagnostic. Photos are matched on Employee ID, PhotoID or
    # WorkdayID (whichever the filename stem happens to be), so report against
    # the images actually in the folder rather than implying every employee
    # should have one. People without a photo get their initials instead.
    if photos:
        matched = sum(1 for e in enriched if e.get("_photo"))
        files = len(set(photos.values()))
        print(f"[ok] Photos matched to employees: {matched} / {len(enriched)} "
              f"(from {files} image file(s); the rest show initials)")
        used = {e["_photo"] for e in enriched if e.get("_photo")}
        if len(used) < files:
            spare = sorted({k for k, v in photos.items() if v not in used and not k.isdigit()}
                           or {k for k, v in photos.items() if v not in used})
            print(f"[note] {files - len(used)} image(s) match nobody on the roster: {spare[:8]}")

    # Stats
    exps = [e.get(EXP_TOTAL) for e in enriched if isinstance(e.get(EXP_TOTAL), (int, float))]
    avg_exp = round(sum(exps) / len(exps), 1) if exps else 0
    ous   = sorted({str(e.get("OU", "")).strip()       for e in enriched if str(e.get("OU", "")).strip()})
    locs  = sorted({str(e.get("Location", "")).strip() for e in enriched if str(e.get("Location", "")).strip()})
    roles = sorted({str(e.get("Role", "")).strip()     for e in enriched if str(e.get("Role", "")).strip()})

    def opts(values):
        return "\n".join(f'<option value="{v}">{v}</option>' for v in values)

    # Day with no leading zero ("31 May 2026" not "31 May 2026, 17:42")
    today = datetime.date.today()
    generated_str = f"{today.day} {today.strftime('%B %Y')}"

    template = TEMPLATE.read_text(encoding="utf-8")
    mobile_template = MOBILE_TEMPLATE.read_text(encoding="utf-8") if MOBILE_TEMPLATE.exists() else None

    def render(share, tpl=None):
        """share=True -> restricted build: utilization stripped + full Address redacted.
        tpl overrides the page template (used for the mobile build)."""
        emp = enriched
        util = data["utilization"]
        util_jj = data["util_jj"]
        util_am = data["util_am"]
        rates = data["rates"]
        leavers = build_leavers(data["details"], util_jj, util_am)
        if share:
            util = util_jj = util_am = []                  # no utilization anywhere
            leavers = []                                   # and so nobody to add
            # The Rate Analysis tab is removed at runtime in share mode, but the
            # rate card itself was still embedded and readable in the page
            # source. Drop it here so what is not shown is also not shipped.
            rates = []
            emp = [{k: v for k, v in e.items() if k != "Address"} for e in enriched]  # redact full address
        return ((tpl if tpl is not None else template)
            .replace("__EMPLOYEES_JSON__", json.dumps(emp, default=str))
            .replace("__SKILLS_JSON__",    json.dumps(data["skills"], default=str))
            .replace("__UTIL_JSON__",      json.dumps(util, default=str))
            .replace("__UTIL_JJ_JSON__",   json.dumps(util_jj, default=str))
            .replace("__UTIL_AM_JSON__",   json.dumps(util_am, default=str))
            .replace("__RATES_JSON__",     json.dumps(rates, default=str))
            .replace("__LEAVERS_JSON__",   json.dumps(leavers, default=str))
            .replace("__TOTAL__",          str(len(enriched)))
            .replace("__AVG_EXP__",        str(avg_exp))
            .replace("__NUM_OUS__",        str(len(ous)))
            .replace("__NUM_LOCS__",       str(len(locs)))
            .replace("__OU_OPTIONS__",     opts(ous))
            .replace("__ROLE_OPTIONS__",   opts(roles))
            .replace("__LOC_OPTIONS__",    opts(locs))
            .replace("__GENERATED__",      generated_str)
            .replace("__LOGO_URI__",       logo_uri)
            .replace("__EXP_TOTAL__",      EXP_TOTAL)
            .replace("__EXP_PWC__",        EXP_PWC)
            .replace("__SHARE_MODE__",     "true" if share else "false")
            .replace("__PBI_URL__",        args.pbi_url))

    builds = [
        (False, OUTPUT_DIR / "Employee_Dashboard.html",        "MD / full (all data)", None),
        (True,  OUTPUT_DIR / "Employee_Dashboard_Shared.html", "Restricted (no utilization, Address redacted)", None),
    ]
    if mobile_template is not None:
        builds.append((False, OUTPUT_DIR / "Employee_Dashboard_Mobile.html",
                       "Mobile / full (Directory + Skill Finder)", mobile_template))

    print(f"[ok] Employees: {len(enriched)} | Skill rows: {len(data['skills'])} | Util rows: {len(data['utilization'])} | Rate rows: {len(data['rates'])}")
    ids_with_skills = {str(r.get("WorkdayID")) for r in data["skills"] if r.get("WorkdayID") not in (None, "")}
    no_skill_emps = [e for e in enriched if str(e.get("WorkdayID")) not in ids_with_skills]
    if no_skill_emps:
        names = ", ".join(str(e.get("Name", "?")) for e in no_skill_emps[:5])
        more = f" + {len(no_skill_emps) - 5} more" if len(no_skill_emps) > 5 else ""
        print(f"[warn] {len(no_skill_emps)} employee(s) have NO rows in the skills sheet: {names}{more}")
    print(f"[ok] Logo found: {'yes' if logo_uri else 'no'}")
    for share, out_file, label, tpl in builds:
        out_file.write_text(render(share, tpl), encoding="utf-8")
        size_kb = out_file.stat().st_size / 1024
        print(f"[ok] Wrote {out_file.name} ({label}) - {size_kb:.0f} KB")

    # Read-only companion: every figure rendered here in Python, no JavaScript,
    # so it opens in previews that switch scripting off - notably iOS. Optional:
    # if the module is missing the three builds above are unaffected.
    try:
        import build_static
        static_out = OUTPUT_DIR / "Employee_Dashboard_Static.html"
        build_static.build(data, enriched,
                           build_leavers(data["details"], data["util_jj"], data["util_am"]),
                           photos, logo_uri, generated_str, static_out, photo_for)
        print(f"[ok] Wrote {static_out.name} (Static / no JavaScript - opens on iOS) "
              f"- {static_out.stat().st_size / 1024:.0f} KB")
    except ImportError:
        print("[note] build_static.py not found - skipping the no-JavaScript build.")


if __name__ == "__main__":
    main()
